#!/usr/bin/env python
# encoding: utf-8

from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import sys
reload(sys)
sys.setdefaultencoding("utf-8")

import os
import json
from multiprocessing import Process
import random
import uuid
import re
import copy
import math
import urllib
import urllib2
import requests
from openpyxl import load_workbook
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


def parse_resp_data_func(input_str):
    res_dict = {}
    if input_str == "":
        res_dict["general_content"] = ""
        res_dict["responds_source"] = ""
        res_dict["query_intent"] = ""
        res_dict["search_query"] = ""
        res_dict["search_content"] = ""
        res_dict["search_ref"] = ""
        return res_dict
    input_str = input_str.split("event:complete")[1].split("data:")[1]
    data = json.loads(input_str.strip("\n"))
    res_dict["general_content"] = data["data"]["message"]["general_content"]
    res_dict["responds_source"] = data["data"]["message"]["responds_source"]
    res_dict["query_intent"] = data["data"]["message"]["query_intent"]
    res_dict["search_query"] = data["data"]["debug"]["search_query"]
    if data["data"]["debug"]["search_ref"] != "":
        res_dict["search_ref"] = data["data"]["debug"]["search_ref"]
    else:
        res_dict["search_ref"] = data["data"]["debug"]["searchRW_ref"]
    res_dict["search_res_list"] = ['标题'+part.strip() for part in re.split(u'\n\[\d\]标题', res_dict["search_ref"].lstrip("参考来源：")) if len(part) >= 5]
    res_dict["search_res_list"] = res_dict["search_res_list"][: 3] + [""] * (3 - len(res_dict["search_res_list"]))
    return res_dict


def parse_and_process_knowledge_enhancement_eval():
    input_file_name = 'kk_random_qa_query_500.curl_result'
    saved_file_name = 'kk_random_qa_query_500.xlsx'
    #input_file_name = 'kk_random_qa_query_600_vertical.curl_result'
    #saved_file_name = 'kk_random_qa_query_600_vertical.xlsx'

    saved_list = []
    test_num = 0
    trigger_1, trigger_2, diff, knowledge = 0, 0, 0, 0
    no_trigger_1_by_intent = 0
    no_trigger_2_by_intent = 0
    no_trigger_1_by_med = 0
    no_trigger_2_by_med = 0
    intent_ok_1, intent_ok_2 = 0, 0
    input_file = open(input_file_name, "r")
    for line in input_file:
        data = json.loads(line.strip("\n"))
        query = data["query"]
        res_dict1 = parse_resp_data_func(data["resp_data_online"])
        res_dict2 = parse_resp_data_func(data["resp_data_pub"])
        tmp_dict = {"query": query,
                "base_response": res_dict1['general_content'], "base_response_source": res_dict1["responds_source"], "base_query_intent": res_dict1["query_intent"], "base_ref_1": res_dict1["search_res_list"][0], "base_ref_2": res_dict1["search_res_list"][1], "base_ref_3": res_dict1["search_res_list"][2],
                "test_response": res_dict2['general_content'], "test_response_source": res_dict2["responds_source"], "test_query_intent": res_dict2["query_intent"], "test_ref_1": res_dict2["search_res_list"][0], "test_ref_2": res_dict2["search_res_list"][1], "test_ref_3": res_dict2["search_res_list"][2],
                }
        # 修复test用了base的sessid导致query_intent不一致的情况
        tmp_dict['test_query_intent'] = tmp_dict['base_query_intent']
        if tmp_dict['test_query_intent'] != u'知识类':
            tmp_dict['test_ref_1'] = ''
            tmp_dict['test_ref_2'] = ''
            tmp_dict['test_ref_3'] = ''
            #print(tmp_dict['query'])
            #print(tmp_dict['base_query_intent'])
        # 数据统计
        if tmp_dict['base_ref_1'] != "":
            trigger_1 += 1
        else:
            if tmp_dict['base_query_intent'] != u'知识类':
                no_trigger_1_by_intent += 1
            elif tmp_dict['base_response_source'] != 'GeneralLLM':
                no_trigger_1_by_med += 1
        if tmp_dict['base_query_intent'] == u'知识类':
            intent_ok_1 += 1
        if tmp_dict['test_ref_1'] != "":
            trigger_2 += 1
        else:
            if tmp_dict['test_query_intent'] != u'知识类':
                no_trigger_2_by_intent += 1
            elif tmp_dict['test_response_source'] != 'GeneralLLM':
                no_trigger_2_by_med += 1
        if tmp_dict['test_query_intent'] == u'知识类':
            intent_ok_2 += 1
        if tmp_dict['test_query_intent'] != tmp_dict['base_query_intent']:
            print(tmp_dict['query'])
            print(tmp_dict['base_query_intent'])
            print(tmp_dict['test_query_intent'])
        if tmp_dict['base_ref_1'] != "" and tmp_dict['test_ref_1'] != "" and '[SEP]'.join(sorted(res_dict1['search_res_list'])) != '[SEP]'.join(sorted(res_dict2['search_res_list'])):
            diff += 1
        if '[SEP]'.join(sorted(res_dict1['search_res_list'])) == '[SEP]'.join(sorted(res_dict2['search_res_list'])):
            continue
        if tmp_dict['base_response_source'] == 'SearchRef' or tmp_dict['test_response_source'] == 'SearchRef':
            if 'knowledge ' in tmp_dict['test_ref_1']:
                knowledge += 1
                saved_list.append(tmp_dict)
    print(len(saved_list))
    print(trigger_1)
    print(trigger_2)
    print(no_trigger_1_by_intent)
    print(no_trigger_2_by_intent)
    print(no_trigger_1_by_med)
    print(no_trigger_2_by_med)
    print(intent_ok_1)
    print(intent_ok_2)
    print(diff)
    print(knowledge)

    input_file.close()
    print("process [%s] file done." % input_file_name)

    '''
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "sheet1"
    word_list = ['query',
            'base_response', 'base_response_source', 'base_query_intent', 'base_ref_1', 'base_ref_2', 'base_ref_3',
            'test_response', 'test_response_source', 'test_query_intent', 'test_ref_1', 'test_ref_2', 'test_ref_3']
    for i in range(len(word_list)):
        worksheet.cell(1, 1 + i, word_list[i])

    succ_cnt = 0
    for i in range(len(saved_list)):
        try:
            worksheet.cell(i + 2, 1, saved_list[i]["query"])
            worksheet.cell(i + 2, 2, saved_list[i]["base_response"])
            worksheet.cell(i + 2, 3, saved_list[i]["base_response_source"])
            worksheet.cell(i + 2, 4, saved_list[i]["base_query_intent"])
            worksheet.cell(i + 2, 5, saved_list[i]["base_ref_1"])
            worksheet.cell(i + 2, 6, saved_list[i]["base_ref_2"])
            worksheet.cell(i + 2, 7, saved_list[i]["base_ref_3"])
            worksheet.cell(i + 2, 8, saved_list[i]["test_response"])
            worksheet.cell(i + 2, 9, saved_list[i]["test_response_source"])
            worksheet.cell(i + 2,10, saved_list[i]["test_query_intent"])
            worksheet.cell(i + 2,11, saved_list[i]["test_ref_1"])
            worksheet.cell(i + 2,12, saved_list[i]["test_ref_2"])
            worksheet.cell(i + 2,13, saved_list[i]["test_ref_3"])
            succ_cnt += 1
        except:
            pass
    workbook.save(saved_file_name)
    print("saved [%s] file with [%d] done." % (saved_file_name, succ_cnt))
    '''


if __name__ == "__main__":
    parse_and_process_knowledge_enhancement_eval()


